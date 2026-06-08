from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import (
    CHECK_FIELDNAMES,
    DEVICE_TYPE_REFERENCE_FIELDNAMES,
    DEVICE_USAGE_FIELDNAMES,
    INSTRUCTION_REFERENCE_FIELDNAMES,
    IO_FIELDNAMES,
    RAW_FIELDNAMES,
)
from .reference_data import (
    DEVICE_TYPE_REFERENCE_ROWS,
    INSTRUCTION_REFERENCE_ROWS,
)


CHECK_LEVEL_ORDER = {
    "ERROR": 0,
    "WARN": 1,
    "INFO": 2,
}

CHECK_LEVEL_FILLS = {
    "ERROR": PatternFill("solid", fgColor="F4CCCC"),
    "WARN": PatternFill("solid", fgColor="FFF2CC"),
    "INFO": PatternFill("solid", fgColor="D9EAF7"),
}


def check_sort_key(row):
    level = row.get("Level", "")
    category = row.get("Category", "")
    device = row.get("Device", "")
    check_type = row.get("Type", "")

    return (
        CHECK_LEVEL_ORDER.get(level, 99),
        category,
        device,
        check_type,
    )


def count_check_levels(check_rows):
    counts = {
        "ERROR": 0,
        "WARN": 0,
        "INFO": 0,
    }

    for row in check_rows:
        level = row.get("Level", "")

        if level in counts:
            counts[level] += 1

    return counts


def write_sheet(ws, rows, fieldnames):
    ws.append(fieldnames)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([row[field] for field in fieldnames])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def write_check_sheet(ws, check_rows):
    sorted_rows = sorted(check_rows, key=check_sort_key)

    write_sheet(ws, sorted_rows, CHECK_FIELDNAMES)

    for row in ws.iter_rows(min_row=2):
        level = row[0].value
        fill = CHECK_LEVEL_FILLS.get(level)

        if fill is None:
            continue

        for cell in row:
            cell.fill = fill


def write_summary_sheet(
    ws,
    metadata,
    input_rows,
    output_rows,
    check_rows,
    device_usage_rows,
    ladder_source_count,
    comment_count,
):
    check_level_counts = count_check_levels(check_rows)

    ws.append(["Item", "Value"])
    ws.append(["Project name", metadata["ProjectName"]])
    ws.append(["Generated at", metadata["GeneratedAt"]])
    ws.append(["Ladder CSV folder", metadata["LadderCsvFolder"]])
    ws.append(["Device comment file", metadata["DeviceCommentFile"]])
    ws.append(["Output CSV", metadata["OutputCsv"]])
    ws.append(["Output Excel", metadata["OutputExcel"]])
    ws.append(["Ladder CSV files", ladder_source_count])
    ws.append(["Device comments", comment_count])
    ws.append(["Input devices", len(input_rows)])
    ws.append(["Output devices", len(output_rows)])
    ws.append(["Total devices", len(input_rows) + len(output_rows)])
    ws.append(["Device usage rows", len(device_usage_rows)])
    ws.append(["Check items", len(check_rows)])
    ws.append(["Check ERROR", check_level_counts["ERROR"]])
    ws.append(["Check WARN", check_level_counts["WARN"]])
    ws.append(["Check INFO", check_level_counts["INFO"]])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18


def write_io_list_excel(
    output_path,
    metadata,
    input_rows,
    output_rows,
    check_rows,
    device_usage_rows,
    raw_rows,
    ladder_source_count,
    comment_count,
):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    write_summary_sheet(
        summary_ws,
        metadata,
        input_rows,
        output_rows,
        check_rows,
        device_usage_rows,
        ladder_source_count,
        comment_count,
    )

    input_ws = wb.create_sheet("INPUT")
    write_sheet(input_ws, input_rows, IO_FIELDNAMES)

    output_ws = wb.create_sheet("OUTPUT")
    write_sheet(output_ws, output_rows, IO_FIELDNAMES)

    check_ws = wb.create_sheet("CHECK")
    write_check_sheet(check_ws, check_rows)

    device_usage_ws = wb.create_sheet("DEVICE_USAGE")
    write_sheet(device_usage_ws, device_usage_rows, DEVICE_USAGE_FIELDNAMES)

    instruction_reference_ws = wb.create_sheet("INSTRUCTION_REFERENCE")
    write_sheet(
        instruction_reference_ws,
        INSTRUCTION_REFERENCE_ROWS,
        INSTRUCTION_REFERENCE_FIELDNAMES,
    )

    device_type_reference_ws = wb.create_sheet("DEVICE_TYPE_REFERENCE")
    write_sheet(
        device_type_reference_ws,
        DEVICE_TYPE_REFERENCE_ROWS,
        DEVICE_TYPE_REFERENCE_FIELDNAMES,
    )

    raw_ws = wb.create_sheet("RAW_DATA")
    write_sheet(raw_ws, raw_rows, RAW_FIELDNAMES)

    wb.save(output_path)
