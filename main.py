import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SAMPLE_DIR = Path("samples/ladder")
COMMENT_CSV_PATH = Path("samples/device_comments/device_comments.csv")
OUTPUT_DIR = Path("output")
OUTPUT_CSV_PATH = OUTPUT_DIR / "io_list.csv"
OUTPUT_EXCEL_PATH = OUTPUT_DIR / "io_list_report.xlsx"

STEP_INDEX = 0
INSTRUCTION_INDEX = 2
DEVICE_INDEX = 3
NOTE_INDEX = 6

IO_FIELDNAMES = [
    "No",
    "Type",
    "Device",
    "DeviceComment",
    "AddressNo",
    "UsedFiles",
    "Instructions",
    "LogicNotes",
    "Steps",
]

CHECK_FIELDNAMES = [
    "Level",
    "Type",
    "Device",
    "Message",
    "Details",
]


def device_sort_key(device):
    device_type = device[0]
    address = device[1:]

    try:
        address_number = int(address, 16)
    except ValueError:
        address_number = 999999

    return device_type, address_number


def add_device(target, device, instruction, file_name, step):
    if device not in target:
        target[device] = {
            "device": device,
            "files": set(),
            "instructions": set(),
            "logic_notes": set(),
            "steps": set(),
        }

    target[device]["files"].add(file_name)

    if instruction:
        target[device]["instructions"].add(instruction)

    if step:
        target[device]["steps"].add(f"{file_name}:{step}")


def add_logic_note(target, device, note):
    if device in target and note:
        target[device]["logic_notes"].add(note)


def is_note_only_row(row, note):
    if not note:
        return False

    return all(
        not row[index].strip()
        for index in range(min(len(row), NOTE_INDEX))
    )


def read_ladder_csv_files(sample_dir):
    inputs = {}
    outputs = {}
    raw_rows = []

    for sample_path in sorted(sample_dir.glob("*.csv")):
        current_io_type = None
        current_device = None

        with sample_path.open("r", encoding="utf-16", newline="") as file:
            reader = csv.reader(file, delimiter="\t")

            for row_number, row in enumerate(reader, start=1):
                if len(row) <= DEVICE_INDEX:
                    continue

                step = row[STEP_INDEX].strip()
                instruction = row[INSTRUCTION_INDEX].strip()
                device = row[DEVICE_INDEX].strip()
                note = row[NOTE_INDEX].strip() if len(row) > NOTE_INDEX else ""

                if is_note_only_row(row, note) and current_device:
                    if current_io_type == "INPUT":
                        add_logic_note(inputs, current_device, note)
                    elif current_io_type == "OUTPUT":
                        add_logic_note(outputs, current_device, note)
                    continue

                if not device.startswith(("X", "Y")):
                    continue

                raw_rows.append(
                    {
                        "source_file": sample_path.name,
                        "row_number": row_number,
                        "step": step,
                        "instruction": instruction,
                        "device": device,
                        "logic_note": note,
                    }
                )

                if device.startswith("X"):
                    add_device(inputs, device, instruction, sample_path.name, step)
                    current_io_type = "INPUT"
                    current_device = device
                elif device.startswith("Y"):
                    add_device(outputs, device, instruction, sample_path.name, step)
                    current_io_type = "OUTPUT"
                    current_device = device

    return inputs, outputs, raw_rows


def read_device_comments(comment_csv_path):
    comments = {}

    if not comment_csv_path.exists():
        return comments

    with comment_csv_path.open("r", encoding="utf-16", newline="") as file:
        reader = csv.reader(file, delimiter="\t")

        # First row is the project name. Second row is the header.
        next(reader, None)
        next(reader, None)

        for row in reader:
            if len(row) < 2:
                continue

            device = row[0].strip()
            comment = row[1].strip()

            if device and comment:
                comments[device] = comment

    return comments


def build_output_rows(io_type, devices, device_comments):
    rows = []

    for index, device in enumerate(sorted(devices, key=device_sort_key), start=1):
        item = devices[device]
        address_number = device_sort_key(device)[1]

        rows.append(
            {
                "No": index,
                "Type": io_type,
                "Device": item["device"],
                "DeviceComment": device_comments.get(device, ""),
                "AddressNo": address_number,
                "UsedFiles": ",".join(sorted(item["files"])),
                "Instructions": ",".join(sorted(item["instructions"])),
                "LogicNotes": ",".join(sorted(item["logic_notes"])),
                "Steps": ",".join(sorted(item["steps"])),
            }
        )

    return rows


def build_check_rows(inputs, outputs, device_comments):
    check_rows = []

    all_devices = {}
    all_devices.update(inputs)
    all_devices.update(outputs)

    for device in sorted(all_devices, key=device_sort_key):
        item = all_devices[device]

        if device not in device_comments:
            check_rows.append(
                {
                    "Level": "WARN",
                    "Type": "MISSING_DEVICE_COMMENT",
                    "Device": device,
                    "Message": "Device comment is empty.",
                    "Details": ",".join(sorted(item["files"])),
                }
            )

        if not item["logic_notes"]:
            check_rows.append(
                {
                    "Level": "WARN",
                    "Type": "MISSING_LOGIC_NOTE",
                    "Device": device,
                    "Message": "Logic note is empty.",
                    "Details": ",".join(sorted(item["files"])),
                }
            )

        if len(item["files"]) > 1:
            check_rows.append(
                {
                    "Level": "WARN",
                    "Type": "MULTIPLE_USED_FILES",
                    "Device": device,
                    "Message": "Device is used in multiple files.",
                    "Details": ",".join(sorted(item["files"])),
                }
            )

        if len(item["logic_notes"]) > 1:
            check_rows.append(
                {
                    "Level": "WARN",
                    "Type": "MULTIPLE_LOGIC_NOTES",
                    "Device": device,
                    "Message": "Device has multiple logic notes.",
                    "Details": ",".join(sorted(item["logic_notes"])),
                }
            )

    return check_rows


def write_io_list_csv(output_path, rows):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=IO_FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def write_sheet(ws, rows, fieldnames):
    ws.append(fieldnames)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([row[field] for field in fieldnames])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def write_summary_sheet(
    ws,
    input_rows,
    output_rows,
    check_rows,
    ladder_source_count,
    comment_count,
):
    ws.append(["Item", "Value"])
    ws.append(["Ladder CSV files", ladder_source_count])
    ws.append(["Device comments", comment_count])
    ws.append(["Input devices", len(input_rows)])
    ws.append(["Output devices", len(output_rows)])
    ws.append(["Total devices", len(input_rows) + len(output_rows)])
    ws.append(["Check items", len(check_rows)])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18


def write_io_list_excel(
    output_path,
    input_rows,
    output_rows,
    check_rows,
    ladder_source_count,
    comment_count,
):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "SUMMARY"
    write_summary_sheet(
        summary_ws,
        input_rows,
        output_rows,
        check_rows,
        ladder_source_count,
        comment_count,
    )

    input_ws = wb.create_sheet("INPUT")
    write_sheet(input_ws, input_rows, IO_FIELDNAMES)

    output_ws = wb.create_sheet("OUTPUT")
    write_sheet(output_ws, output_rows, IO_FIELDNAMES)

    check_ws = wb.create_sheet("CHECK")
    write_sheet(check_ws, check_rows, CHECK_FIELDNAMES)

    wb.save(output_path)


def print_summary(inputs, outputs, checks, device_comments, csv_path, excel_path):
    print(f"Input devices  : {len(inputs)}")
    print(f"Output devices : {len(outputs)}")
    print(f"Device comments: {len(device_comments)}")
    print(f"Check items    : {len(checks)}")
    print(f"CSV file       : {csv_path}")
    print(f"Excel file     : {excel_path}")


def main():
    inputs, outputs, _ = read_ladder_csv_files(SAMPLE_DIR)
    device_comments = read_device_comments(COMMENT_CSV_PATH)

    input_rows = build_output_rows("INPUT", inputs, device_comments)
    output_rows = build_output_rows("OUTPUT", outputs, device_comments)
    all_rows = input_rows + output_rows
    check_rows = build_check_rows(inputs, outputs, device_comments)
    ladder_source_count = len(list(SAMPLE_DIR.glob("*.csv")))

    write_io_list_csv(OUTPUT_CSV_PATH, all_rows)
    write_io_list_excel(
        OUTPUT_EXCEL_PATH,
        input_rows,
        output_rows,
        check_rows,
        ladder_source_count,
        len(device_comments),
    )

    print_summary(
        inputs,
        outputs,
        check_rows,
        device_comments,
        OUTPUT_CSV_PATH,
        OUTPUT_EXCEL_PATH,
    )


if __name__ == "__main__":
    main()
